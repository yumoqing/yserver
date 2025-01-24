# -*- coding:utf8 -*-
import os
import builtins
import sys
import codecs 
from urllib.parse import quote
import json
import asyncio
from aiohttp import BasicAuth
from traceback import format_exc

import random
import time
import datetime
from openpyxl import Workbook
from tempfile import mktemp

from appPublic.jsonConfig import getConfig
from appPublic.dictObject import DictObject
from appPublic.Singleton import GlobalEnv
from appPublic.argsConvert import ArgsConvert
from appPublic.timeUtils import str2Date,str2Datetime,curDatetime, \
			getCurrentTimeStamp,curDateString, curTimeString, \
			monthfirstday, strdate_add, timestampstr
from appPublic.dataencoder import quotedstr
from appPublic.folderUtils import folderInfo
from appPublic.uniqueID import setNode,getID
from appPublic.unicoding import unicoding,uDict,uObject
from appPublic.Singleton import SingletonDecorator
from appPublic.rc4 import password
from appPublic.registerfunction import RegisterFunction
from appPublic.httpclient import HttpClient

from sqlor.dbpools import DBPools,runSQL,runSQLPaging
from sqlor.filter import DBFilter, default_filterjson
from aiohttp.web import StreamResponse

from .xlsxData import XLSXData
from .uriop import URIOp
from .error import Success, Error, NeedLogin, NoPermission
from .filetest import current_fileno
from .filestorage import FileStorage
from .serverenv import ServerEnv

def basic_auth_headers(user, passwd):
	ba = BasicAuth(login=user, password=passwd)
	return {
		"Authorization":ba.encode()
	}

async def stream_response(request, async_data_generator, content_type='text/html'):
	res = StreamResponse()
	if content_type:
		res.content_type = content_type
	await res.prepare(request)
	async for d in async_data_generator():
		if isinstance(d, bytes):
			await res.write(d)
		elif isinstance(d, str):
			await res.write(d.encode('utf-8'))
		else:
			d = json.dumps(d)
			await res.write(d.encode('utf-8'))
		await res.drain()
	await res.write_eof()
	return res

def data2xlsx(rows,headers=None):
	wb = Workbook()
	ws = wb.active

	i = 1
	if headers is not None:
		for j in range(len(headers)):
			v = headers[j].title if headers[j].get('title',False) else headers[j].name
			ws.cell(column=j+1,row=i,value=v)
		i += 1
	for r in rows:
		for j in range(len(r)):
			v = r[headers[j].name]
			ws.cell(column=j+1,row=i,value=v)
		i += 1
	name = mktemp(suffix='.xlsx')
	wb.save(filename = name)
	wb.close()
	return name
	
async def save_file(str_or_bytes, filename):
	fs = FileStorage()
	r = await fs.save(filename, str_or_bytes)
	return r

def webpath(path):
	fs = FileStorage()
	return fs.webpath(path)

def realpath(path):
	fs = FileStorage()
	return fs.realPath(path)

class FileOutZone(Exception):
	def __init__(self,fp,*args,**kwargs):
		super(FileOutZone,self).__init__(*args,**kwargs)
		self.openfilename = fp
	
	def __str__(self):
		return self.openfilename + ': not allowed to open'
		
def get_config_value(kstr):
	keys = kstr.split('.')
	config = getConfig()
	if config is None:
		raise Exception('getConfig() error')
	for k in keys:
		config = config.get(k)
		if not config:
			return None
	return config

def get_definition(k):
	k = f'definitions.{k}'
	return get_config_value(k)

def openfile(url,m):
	fp = abspath(url)
	if fp is None:
		print(f'openfile({url},{m}),url is not match a file')
		raise Exception('url can not mathc a file')
	config = getConfig()
	paths = [ os.path.abspath(p) for p in config.website.paths ]
	fs = config.get('allow_folders',[])
	fs = [ os.path.abspath(i) for i in fs + paths ]
	r = False
	for f in fs:
		if fp.startswith(f):
			r = True
			break
	if not r:
		raise FileOutZone(fp)
	return open(fp,m)
	
def isNone(a):
	return a is None

def abspath(path):
	config = getConfig()
	paths = [ os.path.abspath(p) for p in config.website.paths ]
	for root in paths:
		p = root + path
		if os.path.exists(root+path):
			return p
			
	return None

def appname():
	config = getConfig()
	try:
		return config.license.app
	except:
		return "test app"
	
def configValue(ks):
	config = getConfig()
	try:
		a = eval('config' + ks)
		return a
	except:
		return None

def visualcoding():
	return configValue('.website.visualcoding');

def file_download(request,path,name,coding='utf8'):
	f = openfile(path,'rb')
	b = f.read()
	f.close()
	fname = quote(name).encode(coding)
	hah = b"attachment; filename=" + fname
	# print('file head=',hah.decode(coding))
	request.setHeader(b'Content-Disposition',hah)
	request.setHeader(b'Expires',0)
	request.setHeader(b'Cache-Control',b'must-revalidate, post-check=0, pre-check=0')
	request.setHeader(b'Content-Transfer-Encoding',b'binary')
	request.setHeader(b'Pragma',b'public')
	request.setHeader(b'Content-Length',len(b))
	request.write(b)
	request.finish()
	
def paramify(data, ns):
	ac = ArgsConvert('${', '}$')
	return ac.convert(data, ns)

def initEnv():
	pool = DBPools()
	g = ServerEnv()
	set_builtins()
	g.paramify = paramify
	g.configValue = configValue
	g.visualcoding = visualcoding
	g.uriop = URIOp
	g.isNone = isNone
	g.json = json
	g.ArgsConvert = ArgsConvert
	g.time = time
	g.curDateString = curDateString
	g.curTimeString = curTimeString
	g.datetime = datetime
	g.random = random
	g.str2date = str2Date
	g.str2datetime = str2Datetime
	g.timestampstr = timestampstr
	g.monthfirstday = monthfirstday
	g.curDatetime = curDatetime
	g.strdate_add = strdate_add
	g.uObject = uObject
	g.uuid = getID
	g.runSQL = runSQL
	g.runSQLPaging = runSQLPaging
	g.runSQLIterator = pool.runSQL
	g.runSQLResultFields = pool.runSQLResultFields
	g.getTables = pool.getTables
	g.getTableFields = pool.getTableFields
	g.getTablePrimaryKey = pool.getTablePrimaryKey
	g.getTableForignKeys = pool.getTableForignKeys
	g.folderInfo = folderInfo
	g.abspath = abspath
	g.data2xlsx = data2xlsx
	g.xlsxdata = XLSXData
	g.openfile = openfile
	g.DBPools = DBPools
	g.DBFilter = DBFilter
	g.default_filterjson = default_filterjson
	g.Error = Error
	g.Success = Success
	g.NeedLogin = NeedLogin
	g.NoPermission = NoPermission
	g.password_encode = password
	g.current_fileno = current_fileno
	g.get_config_value = get_config_value
	g.get_definition = get_definition
	g.DictObject = DictObject
	g.async_sleep = asyncio.sleep
	g.quotedstr = quotedstr
	g.save_file = save_file
	g.realpath = realpath
	g.format_exc = format_exc
	g.basic_auth_headers = basic_auth_headers
	g.HttpClient = HttpClient
	g.rfexe = RegisterFunction().exe
	g.stream_response = stream_response
	g.webpath = webpath

def set_builtins():
	all_builtins = [ i for i in dir(builtins) if not i.startswith('_')]
	g = ServerEnv()
	gg = globals()
	for l in all_builtins:
		exec(f'g["{l}"] = {l}',{'g':g})
